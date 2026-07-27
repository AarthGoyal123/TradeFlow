"""OpenPyXL-backed workbook loading."""

from collections.abc import Iterator
from pathlib import Path
from typing import Any, cast
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from app.core.errors import WorkbookValidationError
from app.domain.workbooks.models import CellValue, HeaderCell, WorkbookRow, WorksheetHeader


class OpenPyXLWorkbookLoader:
    """Load `.xlsx` workbooks through OpenPyXL."""

    def load(self, workbook_path: Path) -> "OpenPyXLWorkbookDocument":
        """Load a workbook document."""
        if workbook_path.suffix.lower() != ".xlsx":
            raise WorkbookValidationError(
                "Workbook extension is not supported by the OpenPyXL loader",
                details={"path": str(workbook_path), "supported_extensions": [".xlsx"]},
            )
        if not workbook_path.exists():
            raise WorkbookValidationError(
                "Workbook file does not exist",
                details={"path": str(workbook_path)},
            )
        try:
            workbook: Any = load_workbook(workbook_path, read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise WorkbookValidationError(
                "Workbook could not be read",
                details={"path": str(workbook_path)},
            ) from exc
        return OpenPyXLWorkbookDocument(workbook)


class OpenPyXLWorkbookDocument:
    """Workbook abstraction wrapping an OpenPyXL workbook."""

    def __init__(self, workbook: Any) -> None:
        self._workbook = workbook

    @property
    def sheet_names(self) -> tuple[str, ...]:
        """Return workbook sheet names."""
        return tuple(str(name) for name in cast(list[str], self._workbook.sheetnames))

    def first_sheet(self) -> "OpenPyXLWorksheetReader":
        """Return the first worksheet."""
        if not self._workbook.sheetnames:
            raise WorkbookValidationError("Workbook contains no worksheets")
        return OpenPyXLWorksheetReader(self._workbook[self._workbook.sheetnames[0]])

    def sheet_by_name(self, sheet_name: str) -> "OpenPyXLWorksheetReader | None":
        """Return a worksheet by name if present."""
        if sheet_name not in self._workbook.sheetnames:
            return None
        return OpenPyXLWorksheetReader(self._workbook[sheet_name])


class OpenPyXLWorksheetReader:
    """Worksheet reader wrapping an OpenPyXL worksheet."""

    def __init__(self, worksheet: Any) -> None:
        self._worksheet = worksheet

    @property
    def name(self) -> str:
        """Return worksheet name."""
        return str(self._worksheet.title)

    def read_header(self, row_number: int = 1) -> WorksheetHeader:
        """Return the configured header row."""
        for row in self.iter_rows(min_row=row_number):
            if row.row_number == row_number:
                cells = tuple(
                    HeaderCell(column_number=index, value=str(value).strip())
                    for index, value in enumerate(row.values, start=1)
                    if value is not None and str(value).strip()
                )
                return WorksheetHeader(row_number=row_number, cells=cells)
        return WorksheetHeader(row_number=row_number, cells=())

    def iter_rows(self, *, min_row: int = 1) -> Iterator[WorkbookRow]:
        """Yield worksheet rows while preserving row numbers."""
        for offset, row in enumerate(
            self._worksheet.iter_rows(min_row=min_row, values_only=True),
            start=min_row,
        ):
            yield WorkbookRow(
                row_number=offset,
                values=tuple(self._normalize_cell(value) for value in row),
            )

    @staticmethod
    def _normalize_cell(value: object) -> CellValue:
        if value == "":
            return None
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned if cleaned else None
        if isinstance(value, int | float | bool) or value is None:
            return value
        return str(value)
