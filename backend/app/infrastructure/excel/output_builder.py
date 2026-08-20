"""OpenPyXL output workbook builder."""

from openpyxl import Workbook

from app.domain.datasets.models import DatasetRow, IntermediateDataset
from app.domain.outputs.models import OutputArtifact, OutputType
from app.domain.outputs.ports import OutputStorage
from app.domain.rules.models import RoutedRow, RuleExecutionReport
from app.domain.workbooks.models import CellValue


class OpenPyXLOutputWorkbookBuilder:
    """Build TradeFlow output workbooks with OpenPyXL."""

    def build(
        self,
        *,
        job_id: str,
        dataset: IntermediateDataset,
        rule_report: RuleExecutionReport,
        output_storage: OutputStorage,
    ) -> tuple[OutputArtifact, ...]:
        """Build all output workbooks and return generated artifacts."""
        routes = {routed.row_number: routed for routed in rule_report.routed_rows}
        artifacts = (
            self._write_rows(
                job_id=job_id,
                output_type=OutputType.CLEAN_DATA,
                filename="Clean_Data.xlsx",
                route_name="clean",
                routes=routes,
                rows=tuple(
                    row
                    for row in dataset.rows
                    if _route_for(row=row, routes=routes).route == "clean"
                ),
                dataset=dataset,
                output_storage=output_storage,
            ),
            self._write_rows(
                job_id=job_id,
                output_type=OutputType.REMOVED_ROWS,
                filename="Removed_Rows.xlsx",
                route_name="removed",
                routes=routes,
                rows=tuple(
                    row
                    for row in dataset.rows
                    if _route_for(row=row, routes=routes).route == "removed"
                ),
                dataset=dataset,
                output_storage=output_storage,
            ),
            self._write_rows(
                job_id=job_id,
                output_type=OutputType.NEEDS_REVIEW,
                filename="Needs_Review.xlsx",
                route_name="needs_review",
                routes=routes,
                rows=tuple(
                    row
                    for row in dataset.rows
                    if _route_for(row=row, routes=routes).route == "needs_review"
                ),
                dataset=dataset,
                output_storage=output_storage,
            ),
            self._write_report(
                job_id=job_id,
                dataset=dataset,
                rule_report=rule_report,
                output_storage=output_storage,
            ),
        )
        return artifacts

    def _write_rows(
        self,
        *,
        job_id: str,
        output_type: OutputType,
        filename: str,
        route_name: str,
        routes: dict[int, RoutedRow],
        rows: tuple[DatasetRow, ...],
        dataset: IntermediateDataset,
        output_storage: OutputStorage,
    ) -> OutputArtifact:
        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = "Data"
        include_metadata = output_type != OutputType.CLEAN_DATA
        if include_metadata:
            worksheet.append(("source_row_number", "route", "route_reason", *dataset.fields))
        else:
            worksheet.append(dataset.fields)
        for row in rows:
            if include_metadata:
                route = _route_for(row=row, routes=routes, route_name=route_name)
                worksheet.append(
                    (
                        row.source_row_number,
                        route.route,
                        route.reason,
                        *(cell.value for cell in row.cells),
                    )
                )
            else:
                worksheet.append(tuple(cell.value for cell in row.cells))
        import io

        with io.BytesIO() as buffer:
            workbook.save(buffer)
            buffer.seek(0)
            return output_storage.save_output(job_id, output_type, buffer)

    def _write_report(
        self,
        *,
        job_id: str,
        dataset: IntermediateDataset,
        rule_report: RuleExecutionReport,
        output_storage: OutputStorage,
    ) -> OutputArtifact:
        workbook = Workbook()
        summary = workbook.active
        assert summary is not None
        summary.title = "Summary"
        summary.append(("metric", "value"))
        summary.append(("template_id", dataset.template_id))
        summary.append(("row_count", dataset.row_count))
        summary.append(("rules_evaluated", rule_report.rules_evaluated))
        summary.append(("rule_matches", len(rule_report.matches)))
        summary.append(("validation_findings", len(rule_report.validation_findings)))
        routes = workbook.create_sheet("Routes")
        routes.append(("source_row_number", "route", "reason"))
        for routed_row in rule_report.routed_rows:
            routes.append((routed_row.row_number, routed_row.route, routed_row.reason))
        matches = workbook.create_sheet("Rule Matches")
        matches.append(("rule_id", "row_number", "field", "value", "confidence", "message"))
        for match in rule_report.matches:
            matches.append(
                (
                    match.rule_id,
                    match.row_number,
                    match.field,
                    _cell_value(match.original_value),
                    match.confidence,
                    match.message,
                )
            )
        import io

        with io.BytesIO() as buffer:
            workbook.save(buffer)
            buffer.seek(0)
            return output_storage.save_output(job_id, OutputType.PROCESSING_REPORT, buffer)


def _cell_value(value: CellValue) -> CellValue:
    return value


def _route_for(
    *,
    row: DatasetRow,
    routes: dict[int, RoutedRow],
    route_name: str = "clean",
) -> RoutedRow:
    return routes.get(row.source_row_number) or RoutedRow(
        row_number=row.source_row_number,
        route=route_name,
        reason="No review or removal rules matched",
    )
