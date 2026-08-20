from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.api.dependencies import (
    get_job_executor,
    get_job_service,
    get_output_storage,
    get_processing_report_repository,
    get_processing_service,
)
from app.application.jobs.service import JobService
from app.application.processing.cleaning_service import DataCleaningService
from app.application.processing.dataset_builder import IntermediateDatasetBuilder
from app.application.processing.service import ProcessingService
from app.application.processing.stages import ColumnRemovalStage, NormalizationStage
from app.application.rules.service import RuleEvaluationService
from app.application.rules.transformations import RuleTransformationApplier
from app.application.workbooks.column_mapper import TemplateColumnMapper
from app.application.workbooks.validation import WorkbookValidationService
from app.domain.rules.models import RulePackDefinition
from app.infrastructure.excel.openpyxl_loader import OpenPyXLWorkbookLoader
from app.infrastructure.excel.output_builder import OpenPyXLOutputWorkbookBuilder
from app.infrastructure.files.local_outputs import LocalOutputStorage
from app.infrastructure.files.local_uploads import LocalUploadedFileStorage
from app.infrastructure.jobs.local_executor import SynchronousJobExecutor
from app.infrastructure.persistence.sqlite_jobs import SQLiteJobRepository
from app.infrastructure.template_store.filesystem import FileSystemTemplateRepository
from app.main import create_app
from tests.helpers.auth import create_test_user, override_auth


def test_processing_api_uploads_processes_downloads_and_reports(tmp_path) -> None:
    client, _, _, _ = _build_client(tmp_path)

    workbook_path = _create_test_workbook(tmp_path)
    with open(workbook_path, "rb") as f:
        upload_response = client.post(
            "/jobs",
            data={"template_id": "indian_rice_exports"},
            files={
                "file": (
                    "shipment.xlsx",
                    f.read(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert upload_response.status_code == 200
    job_id = upload_response.json()["job_id"]

    process_response = client.post(f"/jobs/{job_id}/process")
    assert process_response.status_code == 200
    process_body = process_response.json()
    assert process_body["job_id"] == job_id
    assert process_body["status"] in ("queued", "processing", "completed")

    get_response = client.get(f"/jobs/{job_id}")
    assert get_response.status_code == 200
    assert get_response.json()["status"] == "completed"

    for out_type in ("accepted", "rejected", "review", "report"):
        download_response = client.get(f"/jobs/{job_id}/outputs/{out_type}")
        assert download_response.status_code == 200
        assert download_response.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment; filename=" in download_response.headers["content-disposition"]

    report_response = client.get(f"/jobs/{job_id}/report")
    assert report_response.status_code == 200
    report_body = report_response.json()
    assert report_body["job_id"] == job_id
    assert report_body["status"] == "completed"
    assert report_body["total_rows"] >= 0
    assert report_body["outputs"]
    for output in report_body["outputs"]:
        assert output["output_type"]
        assert output["filename"]
        assert output["path"]


def test_process_missing_job_returns_404(tmp_path) -> None:
    client, _, _, _ = _build_client(tmp_path)
    response = client.post("/jobs/missing-job/process")
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "job_not_found"


def test_download_missing_output_returns_404(tmp_path) -> None:
    client, _, _, _ = _build_client(tmp_path)
    response = client.get("/jobs/missing-job/outputs/accepted")
    assert response.status_code == 404


def test_download_invalid_output_type_returns_404(tmp_path) -> None:
    client, _, _, _ = _build_client(tmp_path)
    job_id = _upload_workbook_via_api(client)
    response = client.get(f"/jobs/{job_id}/outputs/invalid_type")
    assert response.status_code == 404


def test_report_missing_job_returns_404(tmp_path) -> None:
    client, _, _, _ = _build_client(tmp_path)
    response = client.get("/jobs/missing-job/report")
    assert response.status_code == 404


def test_report_unprocessed_job_returns_error(tmp_path) -> None:
    client, _, _, _ = _build_client(tmp_path)
    response = client.post(
        "/jobs",
        data={"template_id": "indian_rice_exports"},
        files={
            "file": (
                "shipment.xlsx",
                b"placeholder",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = response.json()["job_id"]
    report_response = client.get(f"/jobs/{job_id}/report")
    assert report_response.status_code == 400
    assert report_response.json()["error"]["code"] == "job_not_processed"


def _build_client(tmp_path: Path):
    upload_dir = tmp_path / "uploads"
    output_dir = tmp_path / "outputs"
    db_path = tmp_path / "tradeflow.sqlite"
    job_repository = SQLiteJobRepository(db_path)
    template_repository = FileSystemTemplateRepository(Path("../templates"))
    workbook_loader = OpenPyXLWorkbookLoader()
    uploaded_file_storage = LocalUploadedFileStorage(upload_dir, 50)
    output_storage = LocalOutputStorage(output_dir)

    job_service = JobService(
        job_repository=job_repository,
        template_repository=template_repository,
        uploaded_file_storage=uploaded_file_storage,
        allowed_extensions=(".xlsx", ".xls"),
    )

    processing_service = ProcessingService(
        job_repository=job_repository,
        template_repository=template_repository,
        uploaded_file_storage=uploaded_file_storage,
        workbook_loader=workbook_loader,
        workbook_validation_service=WorkbookValidationService(
            template_repository=template_repository,
            workbook_loader=workbook_loader,
            column_mapper=TemplateColumnMapper(),
        ),
        dataset_builder=IntermediateDatasetBuilder(),
        cleaning_service=DataCleaningService(),
        column_removal_stage=ColumnRemovalStage(),
        normalization_stage=NormalizationStage(),
        rule_evaluation_service=RuleEvaluationService(
            rule_pack_repository=_RulePackRepository(),
        ),
        transformation_applier=RuleTransformationApplier(),
        output_workbook_builder=OpenPyXLOutputWorkbookBuilder(),
        output_storage=output_storage,
        processing_report_repository=job_repository,
    )

    app = create_app()
    app.dependency_overrides[get_job_service] = lambda: job_service
    app.dependency_overrides[get_processing_service] = lambda: processing_service
    app.dependency_overrides[get_output_storage] = lambda: output_storage
    app.dependency_overrides[get_processing_report_repository] = lambda: job_repository
    app.dependency_overrides[get_job_executor] = lambda: SynchronousJobExecutor(processing_service)

    # Override auth to simulate logged in user
    user = create_test_user()
    override_auth(app, user, tenant_id="test_tenant")

    return TestClient(app, base_url="http://testserver/api/v1"), upload_dir, output_dir, db_path


def _upload_workbook_via_api(client: TestClient) -> str:
    tmp = tmp_test_workbook()
    with open(tmp, "rb") as f:
        resp = client.post(
            "/jobs",
            data={"template_id": "indian_rice_exports"},
            files={
                "file": (
                    "shipment.xlsx",
                    f.read(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    tmp.unlink()
    return resp.json()["job_id"]


def _create_test_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "test_shipment.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Shipments"
    ws.append(["Consignee", "Port", "Carrier"])
    ws.append(["ACME EXPORTS", "Mundra", "MAERSK LINE"])
    wb.save(path)
    return path


def tmp_test_workbook() -> Path:
    path = Path.cwd() / "_tmp_test_wb.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Shipments"
    ws.append(["Consignee", "Port", "Carrier"])
    ws.append(["ACME EXPORTS", "Mundra", "MAERSK LINE"])
    wb.save(path)
    return path


class _RulePackRepository:
    def list_rule_packs(self, template_id: str) -> tuple[RulePackDefinition, ...]:
        return ()
