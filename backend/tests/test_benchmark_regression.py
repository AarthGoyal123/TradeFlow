"""Regression tests for the benchmark workbook (1006 ALL EXPORT JULY 25.xlsx)."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.application.workbooks.column_mapper import TemplateColumnMapper
from app.application.workbooks.intelligence_service import WorkbookIntelligenceService
from app.application.workbooks.validation import WorkbookValidationService
from app.domain.jobs.models import CreateJob
from app.domain.templates.models import (
    ColumnMapping,
    ColumnsConfig,
    FieldCleaningRuleSchema,
    OutputConfig,
    OutputFiles,
    PipelineConfig,
    RulePack,
    TemplateConfig,
    TemplateDefinition,
    WorkbookConfig,
)
from app.domain.workbooks.synonyms import GlobalSynonymDictionary, IndustrySynonymDictionary
from app.infrastructure.excel.openpyxl_loader import OpenPyXLWorkbookLoader

BENCHMARK_PATH = Path(__file__).resolve().parent.parent / "samples" / "1006 ALL EXPORT JULY 25.xlsx"

BENCHMARK_HEADERS = [
    "Indian Port",
    "CUSH",
    "Date",
    "IEC",
    "Exporter_Name",
    "Exporter_Address",
    "Exporter_City_State",
    "Exporter_PIN",
    "Consignee & Consignee_Address",
    "PORT_CD",
    "COUNTRY",
    "CHP",
    "RITC",
    "Description",
    "Quantity",
    "UQC",
    "Unit Rate in FC",
    "Currency",
    "FOB",
]

BENCHMARK_SAMPLE_ROWS = [
    [
        "PETRAPOLE (INPTPB)",
        "INPTPB",
        "2025-07-04 00:00:00",
        "0589007971",
        "MAHYCO PRIVATE LIMITED",
        "2ND FLOOR,MANISH CHAMBER,B.N.BLOCK  LOCA",
        "NEW DELHI ",
        "0",
        "YES AGRO SCIENCEHOLDING NO.28/5. ROAD NO",
        "BENAPOLE",
        "BANGLADESH",
        "10",
        "10061010",
        "INDIAN PLANTING SEED",
        "4.44",
        "MTS",
        "1866.57",
        "USD",
        "700299.67",
    ],
    [
        "HYDERABAD ICD (INSNF6)",
        "INSNF6",
        "2025-07-09 00:00:00",
        "AAXCA1074E",
        "ADVANTA ENTERPRISES LIMITED",
        "UNIPHOS HOUSE, MADHU PARK, C.D. MARG, 11",
        "MUMBAI",
        "0",
        "ADVANTA SEEDS PHILIPPINES INCUNIT 1507,1",
        "Manila",
        "PHILIPPINES",
        "10",
        "10061010",
        "HYBRID RICE SEEDS ADV 8112",
        "73215",
        "KGS",
        "2.87",
        "USD",
        "17286760.73",
    ],
]


def test_benchmark_workbook_opens_successfully() -> None:
    loader = OpenPyXLWorkbookLoader()
    doc = loader.load(BENCHMARK_PATH)
    assert "Sheet1" in doc.sheet_names


def test_benchmark_workbook_intelligence_report() -> None:
    service = WorkbookIntelligenceService(
        workbook_loader=OpenPyXLWorkbookLoader(),
        column_mapper=TemplateColumnMapper(),
    )
    report = service.analyze(
        workbook_path=BENCHMARK_PATH,
        template=_template(),
    )

    assert len(report.sheets) >= 1
    assert report.structure.header.detected_row == 1
    assert report.structure.data_sample.column_count == 19
    assert report.structure.data_sample.estimated_data_rows > 19000
    assert len(report.raw_header) == 19

    explanations = {e.field: e for e in report.mapping_explanations}
    assert explanations["consignee_name"].matched is True
    assert explanations["port"].matched is True
    assert explanations["hs_code"].matched is True

    matched_count = sum(1 for e in report.mapping_explanations if e.matched)
    assert matched_count >= 16

    assert report.confidence.overall > 0.9


def test_benchmark_maps_all_expected_fields() -> None:
    service = WorkbookIntelligenceService(
        workbook_loader=OpenPyXLWorkbookLoader(),
        column_mapper=TemplateColumnMapper(),
    )
    report = service.analyze(
        workbook_path=BENCHMARK_PATH,
        template=_template(),
    )

    explanations = {e.field: e for e in report.mapping_explanations}

    assert explanations["consignee_name"].source_header == "Consignee & Consignee_Address"
    assert explanations["port"].source_header == "PORT_CD"
    assert explanations["hs_code"].source_header == "RITC"
    assert explanations["country"].source_header == "COUNTRY"
    assert explanations["exporter_name"].source_header == "Exporter_Name"
    assert explanations["fob"].source_header == "FOB"
    assert explanations["date"].source_header == "Date"


def test_benchmark_semantic_detects_country_and_hs() -> None:
    service = WorkbookIntelligenceService(
        workbook_loader=OpenPyXLWorkbookLoader(),
        column_mapper=TemplateColumnMapper(),
    )
    report = service.analyze(
        workbook_path=BENCHMARK_PATH,
        template=_template(),
    )

    assert report.semantic.has_country_data is True
    assert report.semantic.has_hs_code_data is True


def test_benchmark_full_pipeline() -> None:
    """End-to-end test: upload, validate, build dataset, remove columns,
    normalize, reorder, output."""
    import shutil
    import tempfile
    from uuid import uuid4

    from app.application.processing.cleaning_service import DataCleaningService
    from app.application.processing.dataset_builder import IntermediateDatasetBuilder
    from app.application.processing.service import ProcessingService
    from app.application.processing.stages import ColumnRemovalStage, NormalizationStage
    from app.application.rules.service import RuleEvaluationService
    from app.application.rules.transformations import RuleTransformationApplier
    from app.domain.rules.evaluator import RuleEvaluator
    from app.domain.rules.operators import RuleOperatorRegistry, default_operators
    from app.infrastructure.excel.output_builder import OpenPyXLOutputWorkbookBuilder
    from app.infrastructure.files.local_outputs import LocalOutputStorage
    from app.infrastructure.files.local_uploads import LocalUploadedFileStorage
    from app.infrastructure.persistence.sqlite_jobs import SQLiteJobRepository
    from app.infrastructure.rules.filesystem import FileSystemRulePackRepository
    from app.infrastructure.rules.rapidfuzz_operator import RapidFuzzEqualsOperator
    from app.infrastructure.template_store.filesystem import FileSystemTemplateRepository

    tmp = Path(tempfile.mkdtemp())
    try:
        tmpl_root = tmp / "templates"
        shutil.copytree(
            Path(__file__).resolve().parent.parent.parent / "templates" / "indian_rice_exports",
            tmpl_root / "indian_rice_exports",
        )
        (tmpl_root / "indian_rice_exports" / "rules" / "fuzzy_matches.json").unlink(missing_ok=True)

        upload_dir = tmp / "uploads"
        output_dir = tmp / "outputs"
        upload_dir.mkdir()
        output_dir.mkdir()

        t_repo = FileSystemTemplateRepository(tmpl_root)
        loader = OpenPyXLWorkbookLoader()
        gd = GlobalSynonymDictionary()
        ind = IndustrySynonymDictionary(gd)
        mapper = TemplateColumnMapper(global_dict=gd, industry_dict=ind)
        validation = WorkbookValidationService(
            template_repository=t_repo,
            workbook_loader=loader,
            column_mapper=mapper,
            global_dict=gd,
            industry_dict=ind,
        )
        builder = IntermediateDatasetBuilder()
        cr = ColumnRemovalStage()
        norm = NormalizationStage()
        dc = DataCleaningService()
        registry = RuleOperatorRegistry(operators=(*default_operators(), RapidFuzzEqualsOperator()))
        rp_repo = FileSystemRulePackRepository(template_root=tmpl_root, template_repository=t_repo)
        rs = RuleEvaluationService(
            rule_evaluator=RuleEvaluator(registry), rule_pack_repository=rp_repo
        )
        ta = RuleTransformationApplier()
        ob = OpenPyXLOutputWorkbookBuilder()
        us = LocalUploadedFileStorage(upload_dir, 50)
        os = LocalOutputStorage(output_dir)
        jr = SQLiteJobRepository(tmp / "db.sqlite")

        job_id = str(uuid4())
        src = BENCHMARK_PATH
        stored = f"{job_id}.xlsx"
        shutil.copy(src, upload_dir / stored)
        jr.create_job(
            CreateJob(
                job_id=job_id,
                template_id="indian_rice_exports",
                original_filename=src.name,
                stored_filename=stored,
            )
        )

        ps = ProcessingService(
            job_repository=jr,
            template_repository=t_repo,
            uploaded_file_storage=us,
            workbook_loader=loader,
            workbook_validation_service=validation,
            dataset_builder=builder,
            cleaning_service=dc,
            column_removal_stage=cr,
            normalization_stage=norm,
            rule_evaluation_service=rs,
            transformation_applier=ta,
            output_workbook_builder=ob,
            output_storage=os,
            processing_report_repository=jr,
        )
        result = ps.process_job(job_id)
        assert result.errors == ()
        assert result.summary is not None
        s = result.summary

        from openpyxl import load_workbook

        clean_path = next(o.path for o in s.outputs if o.output_type.value == "clean_data")
        wb = load_workbook(clean_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        assert headers[0] == "exporter_name"
        assert "source_row_number" not in headers
        assert "route" not in headers
        assert "route_reason" not in headers
        assert "indian_port" not in headers
        assert "cush" not in headers
        assert "chp" not in headers
        assert "exporter_pin" not in headers
        assert "iec" not in headers
        assert "ritc" not in headers

        print("ACTUAL HEADERS:", headers)
        assert len(headers) == 13
        assert all(isinstance(h, str) for h in headers)

        assert s.total_rows == 19967
        assert s.clean_rows > 0
        assert s.removed_rows > 0
        assert s.clean_rows + s.removed_rows + s.needs_review_rows == s.total_rows

        assert result.dataset.fields == (
            "exporter_name",
            "exporter_address",
            "exporter_city_state",
            "consignee_name",
            "country",
            "port",
            "date",
            "description",
            "quantity",
            "uqc",
            "unit_rate",
            "currency",
            "fob",
        )

        consignee_values = [
            str(ws.cell(row=r, column=4).value)
            for r in range(2, min(ws.max_row + 1, 20))
            if ws.cell(row=r, column=4).value is not None
        ]
        clean_consignees = [v.strip().lower() for v in consignee_values]
        assert all("to order" not in v for v in clean_consignees)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_benchmark_validation_service_no_crash() -> None:
    service = WorkbookValidationService(
        template_repository=_TemplateRepository(_template()),
        workbook_loader=OpenPyXLWorkbookLoader(),
        column_mapper=TemplateColumnMapper(),
    )
    result = service.validate(
        template_id="indian_rice_exports",
        workbook_path=BENCHMARK_PATH,
    )

    assert result is not None
    assert result.template_id == "indian_rice_exports"


@pytest.fixture
def _tmp_dir() -> Path:
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as d:
        yield Path(d)


def test_benchmark_with_reordered_columns(_tmp_dir) -> None:
    reordered = [
        "PORT_CD",
        "Consignee & Consignee_Address",
        "RITC",
        "Date",
        "COUNTRY",
        "Indian Port",
        "CUSH",
        "IEC",
        "Exporter_Name",
        "Exporter_Address",
        "Exporter_City_State",
        "Exporter_PIN",
        "CHP",
        "Description",
        "Quantity",
        "UQC",
        "Unit Rate in FC",
        "Currency",
        "FOB",
    ]
    workbook_path = _create_workbook_from_benchmark(_tmp_dir, reordered)

    service = WorkbookIntelligenceService(
        workbook_loader=OpenPyXLWorkbookLoader(),
        column_mapper=TemplateColumnMapper(),
    )
    report = service.analyze(workbook_path=workbook_path, template=_template())

    explanations = {e.field: e for e in report.mapping_explanations}
    assert explanations["consignee_name"].matched is True
    assert explanations["consignee_name"].column_number == 2
    assert explanations["port"].matched is True
    assert explanations["port"].column_number == 1
    assert explanations["hs_code"].matched is True
    assert explanations["hs_code"].column_number == 3
    assert explanations["country"].matched is True


def test_benchmark_with_extra_columns(_tmp_dir) -> None:
    extra = ["Notes", "Review Status"] + BENCHMARK_HEADERS + ["Internal ID", "Timestamp"]
    workbook_path = _create_workbook_from_benchmark(_tmp_dir, extra)

    service = WorkbookIntelligenceService(
        workbook_loader=OpenPyXLWorkbookLoader(),
        column_mapper=TemplateColumnMapper(),
    )
    report = service.analyze(workbook_path=workbook_path, template=_template())

    explanations = {e.field: e for e in report.mapping_explanations}
    assert explanations["consignee_name"].matched is True
    assert explanations["port"].matched is True
    assert explanations["hs_code"].matched is True
    assert explanations["exporter_name"].matched is True
    assert explanations["fob"].matched is True


def test_benchmark_with_missing_optional_columns(_tmp_dir) -> None:
    missing = [h for h in BENCHMARK_HEADERS if h != "Quantity"]
    workbook_path = _create_workbook_from_benchmark(_tmp_dir, missing)

    service = WorkbookIntelligenceService(
        workbook_loader=OpenPyXLWorkbookLoader(),
        column_mapper=TemplateColumnMapper(),
    )
    report = service.analyze(workbook_path=workbook_path, template=_template())

    explanations = {e.field: e for e in report.mapping_explanations}
    assert explanations["consignee_name"].matched is True
    assert explanations["port"].matched is True
    assert explanations["hs_code"].matched is True
    assert explanations["quantity"].matched is False


def _create_workbook_from_benchmark(temp_dir: Path, headers: list[str]) -> Path:
    workbook_path = temp_dir / "benchmark_copy.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(headers)
    for row in BENCHMARK_SAMPLE_ROWS:
        worksheet.append(row)
    workbook.save(workbook_path)
    return workbook_path


def _template() -> TemplateDefinition:
    required_fields = [
        ColumnMapping(
            field="consignee_name",
            aliases=[
                "Consignee",
                "Importer",
                "Buyer",
                "Notify Party",
                "Consignee Name",
                "Consignee & Consignee_Address",
                "Consignee and Consignee Address",
                "Consignee & Address",
                "Customer",
            ],
        ),
        ColumnMapping(
            field="port",
            aliases=[
                "Port",
                "Destination Port",
                "Discharge Port",
                "POD",
                "PORT_CD",
                "Port Code",
                "Port of Discharge",
                "Unloading Port",
            ],
        ),
        ColumnMapping(
            field="hs_code",
            aliases=[
                "HS Code",
                "HSN Code",
                "RITC",
                "ITC HS Code",
                "Commodity Code",
                "Product Code",
                "Tariff Code",
            ],
        ),
    ]
    optional_fields = [
        ColumnMapping(
            field="shipping_company",
            aliases=[
                "Shipping Line",
                "Carrier",
                "Vessel Operator",
                "Shipping Company",
                "Transporter",
                "Vessel",
            ],
        ),
        ColumnMapping(
            field="indian_port",
            aliases=[
                "Indian Port",
                "Port of Loading",
                "Loading Port",
                "POL",
                "Departure Port",
                "Origin Port",
            ],
        ),
        ColumnMapping(
            field="cush",
            aliases=[
                "CUSH",
                "CUSH Code",
                "Port Code Loading",
                "Loading Port Code",
                "Origin Port Code",
            ],
        ),
        ColumnMapping(
            field="date",
            aliases=[
                "Date",
                "Shipment Date",
                "Export Date",
                "Transaction Date",
                "Invoice Date",
                "Document Date",
            ],
        ),
        ColumnMapping(
            field="iec",
            aliases=[
                "IEC",
                "IEC Code",
                "Importer Exporter Code",
                "IE Code",
                "Import Export Code",
                "Exporter IEC",
            ],
        ),
        ColumnMapping(
            field="exporter_name",
            aliases=[
                "Exporter Name",
                "Exporter_Name",
                "Exporter",
                "Shipper",
                "Exporting Company",
                "Seller",
                "Supplier",
            ],
        ),
        ColumnMapping(
            field="exporter_address",
            aliases=[
                "Exporter Address",
                "Exporter_Address",
                "Shipper Address",
                "Supplier Address",
            ],
        ),
        ColumnMapping(
            field="exporter_city_state",
            aliases=[
                "Exporter City State",
                "Exporter_City_State",
                "Exporter City",
                "Shipper City",
                "Origin City",
            ],
        ),
        ColumnMapping(
            field="exporter_pin",
            aliases=[
                "Exporter PIN",
                "Exporter_PIN",
                "Exporter Pincode",
                "Shipper PIN",
                "Shipper Zip",
                "Postal Code",
            ],
        ),
        ColumnMapping(
            field="country",
            aliases=[
                "COUNTRY",
                "Country",
                "Destination Country",
                "Importing Country",
                "Country of Destination",
                "Consignee Country",
            ],
        ),
        ColumnMapping(
            field="chp",
            aliases=[
                "CHP",
                "CHP Code",
                "CHP Rate",
            ],
        ),
        ColumnMapping(
            field="description",
            aliases=[
                "Description",
                "Product Description",
                "Goods Description",
                "Item Description",
                "Commodity Description",
                "Cargo Description",
            ],
        ),
        ColumnMapping(
            field="quantity",
            aliases=[
                "Quantity",
                "Qty",
                "Net Quantity",
                "Shipment Quantity",
                "Gross Quantity",
            ],
        ),
        ColumnMapping(
            field="uqc",
            aliases=[
                "UQC",
                "Unit",
                "Unit of Quantity",
                "Unit Code",
                "UOM",
                "Unit of Measure",
                "Measurement Unit",
            ],
        ),
        ColumnMapping(
            field="unit_rate",
            aliases=[
                "Unit Rate in FC",
                "Unit Rate",
                "Unit Price",
                "Rate per Unit",
                "Unit Value",
                "Price per Unit",
            ],
        ),
        ColumnMapping(
            field="currency",
            aliases=[
                "Currency",
                "FC Currency",
                "Invoice Currency",
                "Transaction Currency",
                "Foreign Currency",
            ],
        ),
        ColumnMapping(
            field="fob",
            aliases=[
                "FOB",
                "FOB Value",
                "FOB Amount",
                "Free on Board",
                "Invoice Value",
                "Total Value",
                "Shipment Value",
            ],
        ),
    ]
    return TemplateDefinition(
        config=TemplateConfig(
            id="indian_rice_exports",
            name="Indian Rice Export Shipments",
            version="0.1.0",
            description="Starter template for Indian rice export shipment data.",
            workbook=WorkbookConfig(sheet_strategy="first_sheet"),
            enabled_modules=[
                "validation",
                "column_removal",
                "normalization",
                "keyword_rules",
                "regex_rules",
                "fuzzy_matching",
                "confidence_scoring",
                "output_generation",
            ],
        ),
        columns=ColumnsConfig(
            required_fields=required_fields,
            optional_fields=optional_fields,
            remove_columns=[],
        ),
        pipeline=PipelineConfig(
            steps=[
                "validation",
                "column_removal",
                "normalization",
                "keyword_rules",
                "regex_rules",
                "fuzzy_matching",
                "confidence_scoring",
                "output_generation",
            ]
        ),
        output=OutputConfig(
            files=OutputFiles(),
            review_threshold=0.75,
            column_order=[
                "exporter_name",
                "exporter_address",
                "exporter_city_state",
                "consignee_name",
                "country",
                "port",
                "date",
                "description",
                "quantity",
                "uqc",
                "unit_rate",
                "currency",
                "fob",
            ],
            cleaning={
                "consignee_name": FieldCleaningRuleSchema(
                    remove_phrases=["to order"],
                    bank_keywords=[
                        "bank",
                        "hsbc",
                        "hdfc",
                        "icici",
                        "sbi",
                        "axis",
                        "yes bank",
                        "idbi",
                        "kotak",
                        "indusind",
                        "pnb",
                        "canara",
                        "union",
                        "bob",
                        "baroda",
                        "citibank",
                        "standard chartered",
                        "dbs",
                    ],
                ),
            },
        ),
        keyword_rules=RulePack(),
        regex_rules=RulePack(),
        fuzzy_matches=RulePack(),
    )


class _TemplateRepository:
    def __init__(self, template: TemplateDefinition) -> None:
        self._template = template

    def list_templates(self) -> list[TemplateDefinition]:
        return [self._template]

    def get_template(self, template_id: str) -> TemplateDefinition:
        return self._template
